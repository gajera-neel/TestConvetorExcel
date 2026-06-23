from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import EXPORT_DIR


SUMMARY_COLUMNS = {
    "Bill Name",
    "Store Name",
    "Vendor",
    "Seller",
    "Invoice Number",
    "Date",
    "Customer",
    "Buyer",
    "Phone",
    "Mobile",
    "Address",
    "GST Number",
    "GST Amount",
    "Tax",
    "Discount",
    "Total",
    "Payment Method",
    "Payment Mode",
    "Due Date",
    "Place Of Supply",
}

ROW_DETAIL_COLUMNS = {"S.No", "Item", "Description", "Qty", "Rate", "GST %", "Amount"}


def _safe_columns(columns: list[str], rows: list[dict]) -> list[str]:
    detected = list(dict.fromkeys(columns or []))
    for row in rows:
        for key in row.keys():
            if key not in detected:
                detected.append(key)
    return detected or ["Data"]


def _has_value(value: object) -> bool:
    return value is not None and str(value).strip() != ""


def _to_decimal(value: object) -> Decimal | None:
    if not _has_value(value):
        return None
    cleaned = str(value).replace(",", "").replace("Rs.", "").replace("INR", "").replace("₹", "").strip()
    try:
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return None


def _sum_column(rows: list[dict], column: str) -> Decimal:
    total = Decimal("0")
    for row in rows:
        value = _to_decimal(row.get(column))
        if value is not None:
            total += value
    return total.quantize(Decimal("0.01"))


def _format_decimal(value: Decimal) -> str:
    return str(value.quantize(Decimal("0.01")))


def _split_summary_and_raw(rows: list[dict], columns: list[str]) -> tuple[dict[str, str], list[str]]:
    summary: dict[str, str] = {}
    raw_columns: list[str] = []

    for column in columns:
        values = [str(row.get(column, "")).strip() for row in rows if _has_value(row.get(column))]
        unique_values = set(values)

        if values and len(unique_values) == 1 and column not in ROW_DETAIL_COLUMNS and (column in SUMMARY_COLUMNS or len(rows) > 1):
            summary[column] = values[0]
        else:
            raw_columns.append(column)

    return summary, raw_columns or columns


def _build_bill_summary(rows: list[dict], summary_values: dict[str, str], raw_columns: list[str]) -> list[tuple[str, str]]:
    summary_items = list(summary_values.items())

    if "Amount" in raw_columns:
        subtotal = _sum_column(rows, "Amount")
        if subtotal:
            summary_items.append(("Calculated Subtotal", _format_decimal(subtotal)))

    if "Qty" in raw_columns:
        quantity = _sum_column(rows, "Qty")
        if quantity:
            summary_items.append(("Calculated Quantity", _format_decimal(quantity)))

    gst = _to_decimal(summary_values.get("GST Amount") or summary_values.get("Tax"))
    discount = _to_decimal(summary_values.get("Discount"))
    subtotal = _sum_column(rows, "Amount") if "Amount" in raw_columns else Decimal("0")
    if subtotal and gst is not None:
        expected_total = subtotal + gst - (discount or Decimal("0"))
        summary_items.append(("Calculated Total Check", _format_decimal(expected_total)))

    summary_items.extend(
        [
            ("Raw Rows", str(len(rows))),
            ("Raw Columns", str(len(raw_columns))),
            ("Generated At", datetime.now().isoformat(timespec="seconds")),
        ]
    )
    return summary_items


def _style_sheet(sheet, header_fill: PatternFill) -> None:
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.fill = header_fill

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        column_letter = get_column_letter(column_cells[0].column)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 50)


def generate_excel_file(rows: list[dict], columns: list[str] | None = None, filename_prefix: str = "export") -> Path:
    safe_rows = rows or [{}]
    safe_columns = _safe_columns(columns or [], safe_rows)
    summary_values, raw_columns = _split_summary_and_raw(safe_rows, safe_columns)
    frame = pd.DataFrame(safe_rows)

    for column in raw_columns:
        if column not in frame.columns:
            frame[column] = ""
    frame = frame[raw_columns]

    filename = f"{filename_prefix}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    output_path = EXPORT_DIR / filename

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        frame.to_excel(writer, sheet_name="Raw Data", index=False)
        workbook = writer.book
        sheet = writer.sheets["Raw Data"]
        header_fill = PatternFill("solid", fgColor="1E293B")
        _style_sheet(sheet, header_fill)

        summary = workbook.create_sheet("Bill Summary")
        summary.append(["Field", "Value"])
        for key, value in _build_bill_summary(safe_rows, summary_values, raw_columns):
            summary.append([key, value])
        _style_sheet(summary, header_fill)

    return output_path
