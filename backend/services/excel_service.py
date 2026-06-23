from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from excel.generator import generate_excel_file
from services.bill_service import list_bills
from services.dashboard_service import build_dashboard


def _style_headers(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="1E293B")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.fill = header_fill


def _auto_width(sheet) -> None:
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 60)


def create_excel_report(db: Session) -> Path:
    history = list_bills(db)
    dashboard = build_dashboard(db)
    workbook = Workbook()

    raw_sheet = workbook.active
    raw_sheet.title = "Raw Data"
    columns = sorted(
        {
            column
            for record in history
            for row in (record.get("rows") or [record.get("fields", {})])
            for column in {*row.keys(), *record.get("fields", {}).keys()}
        }
    )
    raw_sheet.append(["Upload ID", "Filename", "Detected Type", "Row No", *columns])
    for record in history:
        rows = record.get("rows") or [record.get("fields", {})]
        for row_index, row in enumerate(rows, start=1):
            raw_sheet.append(
                [
                    record.get("id", ""),
                    record.get("filename", ""),
                    record.get("detected_type", ""),
                    row_index,
                    *[row.get(column, record.get("fields", {}).get(column, "")) for column in columns],
                ]
            )

    summary_sheet = workbook.create_sheet("Bill Summary")
    summary_sheet.append(
        [
            "Upload ID",
            "Filename",
            "Bill Name",
            "Invoice Number",
            "Date",
            "Vendor",
            "Detected Type",
            "Amount",
            "Tax",
            "Total",
            "Rows",
            "Confidence",
        ]
    )
    for record in history:
        fields = record.get("fields", {})
        summary_sheet.append(
            [
                record.get("id", ""),
                record.get("filename", ""),
                record.get("bill_name") or fields.get("Bill Name") or fields.get("Store Name") or "",
                record.get("invoice_number") or fields.get("Invoice Number") or "",
                record.get("bill_date") or fields.get("Date") or "",
                record.get("vendor") or fields.get("Seller") or fields.get("Vendor") or "",
                record.get("detected_type", ""),
                record.get("amount", ""),
                record.get("tax", ""),
                record.get("total", ""),
                record.get("rows_count") or len(record.get("rows") or []),
                record.get("confidence", ""),
            ]
        )

    metrics_sheet = workbook.create_sheet("Dashboard Metrics")
    metrics_sheet.append(["Metric", "Value"])
    for key, value in dashboard["metrics"].items():
        metrics_sheet.append([key.replace("_", " ").title(), value])

    for sheet in workbook.worksheets:
        _style_headers(sheet)
        _auto_width(sheet)

    output_path = generate_excel_file(
        rows=[row for record in history for row in (record.get("rows") or [record.get("fields", {})])],
        columns=columns,
        filename_prefix="export",
    )
    workbook.save(output_path)
    return output_path


def create_custom_excel(rows: list[dict], columns: list[str] | None = None) -> Path:
    return generate_excel_file(rows=rows, columns=columns, filename_prefix="export")
